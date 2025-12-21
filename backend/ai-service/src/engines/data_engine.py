"""
Data Engine - 数据处理引擎
生成Excel、CSV文件和数据可视化图表
"""
import os
import json
from typing import Dict, Any, Optional, List
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import io
import ollama
from openai import AsyncOpenAI
import matplotlib
matplotlib.use('Agg')  # 无GUI后端
from src.llm.llm_client import get_llm_client


class DataEngine:
    """数据处理引擎"""

    def __init__(self):
        self.llm_provider = os.getenv("LLM_PROVIDER", "ollama")
        self.model_name = os.getenv("LLM_MODEL", "qwen2:7b")
        self.openai_api_key = os.getenv("OPENAI_API_KEY")
        self.openai_base_url = os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1")

        self.client = None
        self.llm_client = None
        self._ready = True

        if self.llm_provider == "openai":
            if self.openai_api_key:
                self.client = AsyncOpenAI(api_key=self.openai_api_key, base_url=self.openai_base_url)
            else:
                self._ready = False
        elif self.llm_provider != "ollama":
            try:
                self.llm_client = get_llm_client()
            except Exception as e:
                print(f"LLM client initialization error: {e}")
                self._ready = False

        # 设置中文字体（用于matplotlib）
        # 使用文泉驿字体（在Docker容器中已安装）
        plt.rcParams['font.sans-serif'] = ['WenQuanYi Zen Hei', 'WenQuanYi Micro Hei', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题

    def is_ready(self) -> bool:
        """检查引擎是否就绪"""
        return self._ready

    async def generate(
        self,
        prompt: str,
        context: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        生成数据项目

        Args:
            prompt: 用户需求描述
            context: 意图识别结果

        Returns:
            {
                "files": [
                    {"path": "data.xlsx", "content": bytes, "type": "excel"},
                    {"path": "chart.png", "content": bytes, "type": "image"}
                ],
                "metadata": {...}
            }
        """
        if not self._ready:
            raise Exception("Data engine not ready")

        try:
            # 提取上下文信息
            entities = context.get("entities", {}) if context else {}
            data_type = entities.get("data_source", "sample")
            task = entities.get("task", "analysis")  # analysis, visualization, report

            # 生成数据规格
            spec = await self._generate_data_spec(prompt, entities)

            # 生成示例数据
            df = await self._generate_sample_data(spec)

            files = []

            # 生成Excel文件
            excel_bytes = self._create_excel(df, spec)
            files.append({
                "path": f"{spec.get('filename', 'data')}.xlsx",
                "content": excel_bytes,
                "type": "excel"
            })

            # 生成CSV文件
            csv_bytes = df.to_csv(index=False).encode('utf-8-sig')
            files.append({
                "path": f"{spec.get('filename', 'data')}.csv",
                "content": csv_bytes,
                "type": "csv"
            })

            # 生成可视化图表
            if spec.get("include_visualization", True):
                chart_bytes = self._create_visualizations(df, spec)
                files.append({
                    "path": f"{spec.get('filename', 'data')}_chart.png",
                    "content": chart_bytes,
                    "type": "image"
                })

            return {
                "files": files,
                "metadata": {
                    "data_type": data_type,
                    "task": task,
                    "spec": spec,
                    "rows": len(df),
                    "columns": list(df.columns)
                }
            }

        except Exception as e:
            print(f"Data generation error: {e}")
            raise

    async def _generate_data_spec(
        self,
        prompt: str,
        entities: Dict[str, Any]
    ) -> Dict[str, Any]:
        """生成数据规格说明"""

        spec_prompt = f"""根据用户需求生成数据规格说明。

用户需求: {prompt}
提取的实体: {json.dumps(entities, ensure_ascii=False)}

请生成包含以下信息的JSON规格:
{{
    "filename": "sales_data",
    "data_type": "sales",
    "columns": [
        {{"name": "日期", "type": "date"}},
        {{"name": "产品", "type": "string"}},
        {{"name": "销售额", "type": "number"}},
        {{"name": "数量", "type": "integer"}}
    ],
    "row_count": 100,
    "chart_types": ["bar", "line", "pie"],
    "include_visualization": true,
    "summary_stats": true
}}

只返回JSON，不要其他文字:
"""

        try:
            if self.llm_provider == "ollama":
                response = ollama.chat(
                    model=self.model_name,
                    messages=[
                        {"role": "system", "content": "You are a data analysis expert. Return valid JSON only."},
                        {"role": "user", "content": spec_prompt}
                    ],
                    options={"temperature": 0.3}
                )
                content = response['message']['content']
            elif self.client is not None:
                response = await self.client.chat.completions.create(
                    model=self.model_name,
                    messages=[
                        {"role": "system", "content": "You are a data analysis expert. Return valid JSON only."},
                        {"role": "user", "content": spec_prompt}
                    ],
                    temperature=0.3
                )
                content = response.choices[0].message.content
            elif self.llm_client is not None:
                content = await self.llm_client.chat(
                    messages=[
                        {"role": "system", "content": "You are a data analysis assistant. Return valid JSON only."},
                        {"role": "user", "content": spec_prompt}
                    ],
                    temperature=0.3,
                    max_tokens=1024
                )
            else:
                raise Exception("LLM client not initialized")
            spec = json.loads(content)
            return spec

        except Exception as e:
            print(f"Data spec generation error: {e}")
            # 返回默认规格
            return {
                "filename": "data",
                "data_type": "sample",
                "columns": [
                    {"name": "项目", "type": "string"},
                    {"name": "数值", "type": "number"}
                ],
                "row_count": 10,
                "chart_types": ["bar"],
                "include_visualization": True,
                "summary_stats": True
            }

    async def _generate_sample_data(self, spec: Dict[str, Any]) -> pd.DataFrame:
        """生成示例数据"""

        columns = spec.get("columns", [])
        row_count = spec.get("row_count", 10)

        # 根据列定义生成数据
        data = {}

        for col in columns:
            col_name = col.get("name", "Column")
            col_type = col.get("type", "string")

            if col_type == "date":
                data[col_name] = pd.date_range(start='2024-01-01', periods=row_count, freq='D')
            elif col_type == "number":
                import numpy as np
                data[col_name] = np.random.randint(100, 1000, size=row_count)
            elif col_type == "integer":
                import numpy as np
                data[col_name] = np.random.randint(1, 100, size=row_count)
            elif col_type == "string":
                # 生成示例字符串
                data[col_name] = [f"项目{i+1}" for i in range(row_count)]
            else:
                data[col_name] = [f"Value{i+1}" for i in range(row_count)]

        df = pd.DataFrame(data)
        return df

    def _create_excel(self, df: pd.DataFrame, spec: Dict[str, Any]) -> bytes:
        """创建Excel文件"""

        buffer = io.BytesIO()

        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # 写入数据表
            df.to_excel(writer, sheet_name='数据', index=False)

            # 获取workbook和worksheet
            workbook = writer.book
            worksheet = writer.sheets['数据']

            # 设置表头样式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 自动调整列宽
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # 如果需要统计信息
            if spec.get("summary_stats", True):
                summary_sheet = workbook.create_sheet("统计摘要")

                summary_sheet['A1'] = "统计摘要"
                summary_sheet['A1'].font = Font(size=14, bold=True)

                # 添加基本统计
                summary_sheet['A3'] = "总行数:"
                summary_sheet['B3'] = len(df)

                summary_sheet['A4'] = "总列数:"
                summary_sheet['B4'] = len(df.columns)

                # 数值列统计
                numeric_cols = df.select_dtypes(include=['number']).columns
                if len(numeric_cols) > 0:
                    summary_sheet['A6'] = "数值列统计"
                    summary_sheet['A6'].font = Font(bold=True)

                    row = 7
                    for col in numeric_cols:
                        summary_sheet[f'A{row}'] = col
                        summary_sheet[f'B{row}'] = f"平均: {df[col].mean():.2f}"
                        summary_sheet[f'C{row}'] = f"总和: {df[col].sum():.2f}"
                        row += 1

        buffer.seek(0)
        return buffer.getvalue()

    def _create_visualizations(self, df: pd.DataFrame, spec: Dict[str, Any]) -> bytes:
        """创建数据可视化图表"""

        chart_types = spec.get("chart_types", ["bar"])
        num_charts = len(chart_types)

        # 创建子图
        fig, axes = plt.subplots(1, min(num_charts, 3), figsize=(15, 5))
        if num_charts == 1:
            axes = [axes]

        # 获取数值列
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        string_cols = df.select_dtypes(include=['object']).columns.tolist()

        if len(numeric_cols) == 0:
            # 没有数值列，返回空图
            fig.text(0.5, 0.5, '无数值数据可可视化', ha='center', va='center', fontsize=16)
        else:
            for idx, chart_type in enumerate(chart_types[:min(num_charts, 3)]):
                ax = axes[idx] if num_charts > 1 else axes[0]

                if chart_type == "bar":
                    # 柱状图
                    if len(string_cols) > 0 and len(numeric_cols) > 0:
                        x_col = string_cols[0]
                        y_col = numeric_cols[0]
                        # 限制显示前10个
                        plot_df = df.head(10)
                        ax.bar(plot_df[x_col], plot_df[y_col], color='steelblue')
                        ax.set_xlabel(x_col)
                        ax.set_ylabel(y_col)
                        ax.set_title(f'{y_col} 柱状图')
                        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')
                    else:
                        df[numeric_cols[0]].head(10).plot(kind='bar', ax=ax, color='steelblue')
                        ax.set_title(f'{numeric_cols[0]} 柱状图')

                elif chart_type == "line":
                    # 折线图
                    if len(numeric_cols) > 0:
                        df[numeric_cols[0]].plot(kind='line', ax=ax, color='green', marker='o')
                        ax.set_title(f'{numeric_cols[0]} 趋势图')
                        ax.set_xlabel('索引')
                        ax.set_ylabel(numeric_cols[0])

                elif chart_type == "pie":
                    # 饼图
                    if len(string_cols) > 0 and len(numeric_cols) > 0:
                        x_col = string_cols[0]
                        y_col = numeric_cols[0]
                        # 聚合数据（前5个）
                        plot_df = df.groupby(x_col)[y_col].sum().head(5)
                        ax.pie(plot_df.values, labels=plot_df.index, autopct='%1.1f%%', startangle=90)
                        ax.set_title(f'{y_col} 分布')
                    else:
                        # 使用数值列前5个
                        plot_data = df[numeric_cols[0]].head(5)
                        ax.pie(plot_data, labels=df.index[:5], autopct='%1.1f%%')
                        ax.set_title(f'{numeric_cols[0]} 分布')

        plt.tight_layout()

        # 保存到字节流
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
        plt.close(fig)
        buffer.seek(0)
        return buffer.getvalue()
