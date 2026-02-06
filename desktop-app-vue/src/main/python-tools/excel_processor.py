#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel处理工具
使用openpyxl和pandas处理Excel文件
"""

import sys
import json
import os
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    import pandas as pd
except ImportError as e:
    print(json.dumps({
        'success': False,
        'error': f'缺少依赖包: {str(e)}。请运行: pip install openpyxl pandas'
    }, ensure_ascii=False))
    sys.exit(1)

class ExcelProcessor:
    """Excel处理器"""

    def __init__(self):
        self.workbook = None

    def create_workbook(self, params):
        """
        创建Excel工作簿

        参数:
        - title: 工作簿标题
        - sheets: 工作表列表 [{ name, data, chart }]
        - output_path: 输出路径
        - template: 模板类型 (sales/financial/data_analysis)
        """
        title = params.get('title', '未命名工作簿')
        sheets = params.get('sheets', [])
        output_path = params.get('output_path')
        template = params.get('template', 'basic')

        # 创建工作簿
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # 删除默认工作表

        # 根据模板类型创建
        if template == 'sales':
            self._create_sales_report(title, sheets)
        elif template == 'financial':
            self._create_financial_report(title, sheets)
        elif template == 'data_analysis':
            self._create_data_analysis(title, sheets)
        else:
            self._create_basic_workbook(title, sheets)

        # 保存
        if not output_path:
            output_path = f"{title}.xlsx"

        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)

        self.workbook.save(output_path)

        return {
            'success': True,
            'output_path': output_path,
            'file_size': os.path.getsize(output_path),
            'sheets_count': len(self.workbook.sheetnames),
            'created_at': datetime.now().isoformat()
        }

    def _create_basic_workbook(self, title, sheets):
        """创建基础工作簿"""
        if not sheets:
            # 默认创建一个空白工作表
            ws = self.workbook.create_sheet(title='Sheet1')
            ws['A1'] = title
            ws['A1'].font = Font(size=14, bold=True)
            return

        for sheet_data in sheets:
            sheet_name = sheet_data.get('name', 'Sheet')
            data = sheet_data.get('data', [])

            ws = self.workbook.create_sheet(title=sheet_name)

            # 填充数据
            if isinstance(data, list) and len(data) > 0:
                # 如果是列表，逐行填充
                for row_idx, row_data in enumerate(data, start=1):
                    if isinstance(row_data, list):
                        for col_idx, cell_value in enumerate(row_data, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    elif isinstance(row_data, dict):
                        # 如果是字典，第一行作为标题
                        if row_idx == 1:
                            for col_idx, key in enumerate(row_data.keys(), start=1):
                                cell = ws.cell(row=1, column=col_idx, value=key)
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
                        for col_idx, value in enumerate(row_data.values(), start=1):
                            ws.cell(row=row_idx + 1, column=col_idx, value=value)

            # 添加图表（如果指定）
            if 'chart' in sheet_data:
                self._add_chart(ws, sheet_data['chart'])

    def _create_sales_report(self, title, sheets):
        """创建销售报表"""
        ws = self.workbook.create_sheet(title='销售概览')

        # 标题
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # 示例数据结构
        ws['A3'] = '月份'
        ws['B3'] = '销售额'
        ws['C3'] = '成本'
        ws['D3'] = '利润'
        ws['E3'] = '利润率'

        # 设置标题行样式
        for col in range(1, 6):
            cell = ws.cell(row=3, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # 如果有数据，填充
        if sheets and len(sheets) > 0:
            data = sheets[0].get('data', [])
            for row_idx, row_data in enumerate(data, start=4):
                if isinstance(row_data, list):
                    for col_idx, value in enumerate(row_data, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

        # 添加条形图
        chart = BarChart()
        chart.title = '月度销售趋势'
        chart.x_axis.title = '月份'
        chart.y_axis.title = '金额'

        data_ref = Reference(ws, min_col=2, min_row=3, max_row=ws.max_row, max_col=4)
        cats = Reference(ws, min_col=1, min_row=4, max_row=ws.max_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, 'G3')

    def _create_financial_report(self, title, sheets):
        """创建财务报表"""
        ws = self.workbook.create_sheet(title='财务报表')

        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        # 财务报表基本结构
        headers = ['科目', '本期金额', '上期金额', '增长率']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

        # 填充数据
        if sheets and len(sheets) > 0:
            data = sheets[0].get('data', [])
            for row_idx, row_data in enumerate(data, start=4):
                if isinstance(row_data, list):
                    for col_idx, value in enumerate(row_data, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

    def _create_data_analysis(self, title, sheets):
        """创建数据分析表"""
        ws = self.workbook.create_sheet(title='数据分析')

        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True)

        # 统计摘要
        ws['A3'] = '数据统计'
        ws['A3'].font = Font(bold=True, size=12)

        stats_labels = ['总计', '平均值', '最大值', '最小值', '标准差']
        for idx, label in enumerate(stats_labels, start=4):
            ws[f'A{idx}'] = label

        # 如果有数据，填充
        if sheets and len(sheets) > 0:
            data = sheets[0].get('data', [])
            if data and isinstance(data, list):
                # 使用pandas计算统计量
                try:
                    df = pd.DataFrame(data[1:], columns=data[0])
                    numeric_cols = df.select_dtypes(include=['number']).columns

                    for col_idx, col in enumerate(numeric_cols, start=2):
                        ws.cell(row=3, column=col_idx, value=col)
                        ws.cell(row=4, column=col_idx, value=df[col].sum())
                        ws.cell(row=5, column=col_idx, value=df[col].mean())
                        ws.cell(row=6, column=col_idx, value=df[col].max())
                        ws.cell(row=7, column=col_idx, value=df[col].min())
                        ws.cell(row=8, column=col_idx, value=df[col].std())
                except Exception as e:
                    print(f"统计计算失败: {e}", file=sys.stderr)

    def _add_chart(self, worksheet, chart_config):
        """添加图表"""
        chart_type = chart_config.get('type', 'bar')
        data_range = chart_config.get('data_range', 'B2:C10')
        position = chart_config.get('position', 'E2')

        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'pie':
            chart = PieChart()
        elif chart_type == 'line':
            chart = LineChart()
        else:
            return

        chart.title = chart_config.get('title', '图表')

        # 解析数据范围 (例如 "B2:C10")
        try:
            # 获取数据引用
            min_col = ord(data_range[0].upper()) - ord('A') + 1
            max_col = ord(data_range[3].upper()) - ord('A') + 1
            min_row = int(data_range[1])
            max_row = int(data_range[4:])

            data = Reference(worksheet, min_col=min_col, min_row=min_row,
                           max_col=max_col, max_row=max_row)
            chart.add_data(data, titles_from_data=True)

            # 类别标签
            cats = Reference(worksheet, min_col=1, min_row=min_row+1, max_row=max_row)
            chart.set_categories(cats)
        except Exception:
            pass  # 如果解析失败，继续使用默认设置

        # 图表尺寸
        chart.width = chart_config.get('width', 15)
        chart.height = chart_config.get('height', 10)

        # 图表样式
        if chart_config.get('style'):
            chart.style = chart_config.get('style')

        # 图例位置
        if chart_config.get('legend_position'):
            chart.legend.position = chart_config.get('legend_position')

        # Y轴标题
        if chart_config.get('y_axis_title') and hasattr(chart, 'y_axis'):
            chart.y_axis.title = chart_config.get('y_axis_title')

        # X轴标题
        if chart_config.get('x_axis_title') and hasattr(chart, 'x_axis'):
            chart.x_axis.title = chart_config.get('x_axis_title')

        worksheet.add_chart(chart, position)

def main():
    """主函数"""
    try:
        if len(sys.argv) < 2:
            raise ValueError('缺少参数')

        args = json.loads(sys.argv[1])

        processor = ExcelProcessor()

        # 执行操作
        operation = args.get('operation', 'create')

        if operation == 'create':
            result = processor.create_workbook(args)
        else:
            raise ValueError(f"不支持的操作: {operation}")

        print(json.dumps(result, ensure_ascii=False, indent=2))

    except Exception as e:
        error_result = {
            'success': False,
            'error': str(e),
            'type': type(e).__name__
        }
        print(json.dumps(error_result, ensure_ascii=False))
        sys.exit(1)

if __name__ == '__main__':
    main()
